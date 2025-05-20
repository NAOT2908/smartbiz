from odoo.tools.float_utils import float_compare, float_is_zero, float_round
# -*- coding: utf-8 -*-

from odoo.osv import expression
from odoo import models, fields, api, exceptions,_, tools
import os
import base64,pytz,logging
from datetime import datetime, timedelta
import datetime as date_time
import random
from odoo.exceptions import UserError, ValidationError
_logger = logging.getLogger(__name__)
from io import BytesIO
import xlsxwriter
from openpyxl import load_workbook
class SmartBiz_Stock_StockReport(models.Model):
    _inherit = ['smartbiz_stock.stock_report']
    from_date = fields.Date(store='True')


    def action__nxt__ten_nvl____sl_ht(self):
        file_name = 'Bao cao NXT theo ten NVL va SL he thong 2 sheet.xlsx'    
        workbook = self.load_excel(file_name)
        
        worksheet_met = workbook['Don vi chinh']   # Sheet 1
        worksheet_cuon = workbook['Don vi quy doi'] # Sheet 2

        from_date = self.from_date
        to_date = self.to_date + timedelta(days=1)

        transaction_data = self._get_stock_data(from_date, to_date)

        detailed_data = []
        for (product_id, lot_id, warehouse_id), data in transaction_data.items():
            product = self.env['product.product'].browse(product_id)
            lot = self.env['stock.lot'].browse(lot_id) if lot_id else False
            warehouse = self.env['stock.warehouse'].browse(warehouse_id) if warehouse_id else False
            lot_values = self._get_values_from_split(lot.name if lot else '', ':', 4)
            product_values = self._get_values_from_split(product.name if product else '', ':', 2)
            warehouse_name = warehouse.name if warehouse else ''
            
            detailed_data.append({
                'warehouse': warehouse_name,
                'he_hang': lot_values[0],
                'order': lot_values[1],
                'mass_hanbai': lot_values[2],
                'bom_add': lot_values[3],
                'product': product_values[0],
                'color': product_values[1],
                'ma_hq': product.customs_code,
                'hsqd': product.convert_factor or 1.0,
                'rate': product.convert_rate,
                'begin_quantity': data['begin_quantity'],
                'in_purchase': data['in_purchase'],
                'in_production': data['in_production'],
                'in_adjust': data['in_adjust'],
                'in_transfer': data['in_transfer'],
                'in_other': data['in_other'],
                'in_total': data['in_total'],
                'out_production': data['out_production'],
                'out_supplier': data['out_supplier'],
                'out_huy': data['out_huy'],
                'out_adjust': data['out_adjust'],
                'out_transfer': data['out_transfer'],
                'out_other': data['out_other'],
                'out_total': data['out_total'],
                'end_quantity': data['end_quantity'],
            })

        detailed_data = sorted(detailed_data, key=lambda x: x['warehouse'])

        # Sheet Don vi met
        worksheet_met['A3'] = "Range Date: " + from_date.strftime('%d/%m/%Y') + " - " + to_date.strftime('%d/%m/%Y')
        row_met = 6
        for data in detailed_data:
            keys = list(data.keys())
            for index, key in enumerate(keys, start=1):
                worksheet_met.cell(row_met, index, data[key])
            row_met += 1

        # Sheet Don vi cuon
        worksheet_cuon['A3'] = "Range Date: " + from_date.strftime('%d/%m/%Y') + " - " + to_date.strftime('%d/%m/%Y')
        row_cuon = 6
        quantity_keys = [
            'begin_quantity', 'in_purchase', 'in_production', 'in_adjust', 'in_transfer', 'in_other', 'in_total',
            'out_production', 'out_supplier', 'out_huy', 'out_adjust', 'out_transfer', 'out_other', 'out_total', 'end_quantity'
        ]
        for data in detailed_data:
            keys = list(data.keys())
            for index, key in enumerate(keys, start=1):
                value = data[key]

                if key in quantity_keys:
                    original = value or 0
                    rate_raw = data.get('rate')

                    try:
                        rate = float(rate_raw)
                        if rate and rate != 1.0:
                            converted = original / rate
                            cell_value = f"{original} ; {converted}"
                        else:
                            cell_value = original
                    except (ValueError, TypeError):
                        cell_value = original
                else:
                    cell_value = value

                worksheet_cuon.cell(row_cuon, index, cell_value)
            row_cuon += 1

        return self.save_excel(workbook, file_name)

        
        
    def action__nxt__ten_nvl___sl_qd_hq(self):
        file_name = 'Bao cao NXT theo ten NVL va SL quy doi HQ.xlsx'
        workbook = self.load_excel(file_name)
        worksheet = workbook.active
        from_date = self.from_date
        to_date = self.to_date + timedelta(days=1)

        transaction_data = self._get_stock_data(from_date, to_date)

        # Biến đổi dữ liệu để dễ dàng hiển thị hoặc trả về
        detailed_data = []
        for (product_id, lot_id, warehouse_id), data in transaction_data.items():
            product = self.env['product.product'].browse(product_id)
            lot = self.env['stock.lot'].browse(lot_id) if lot_id else False
            warehouse = self.env['stock.warehouse'].browse(warehouse_id) if warehouse_id else False
            lot_values = self._get_values_from_split(lot.name if lot else '', ':', 4)
            product_values = self._get_values_from_split(product.name if product else '', ':', 2)
            warehouse_name = warehouse.name if warehouse else ''
            detailed_data.append({
                'warehouse': warehouse_name,
                'he_hang': lot_values[0],
                'order': lot_values[1],
                'mass_hanbai': lot_values[2],
                'bom_add': lot_values[3],
                'product': product_values[0],
                'color': product_values[1],
                'ma_hq': product.customs_code,
                'hsqd': product.convert_factor,
                'begin_quantity': data['begin_quantity'] * product.convert_factor if product.convert_factor != 0 else data['begin_quantity'],
                'in_purchase': data['in_purchase'] * product.convert_factor if product.convert_factor != 0 else data['in_purchase'],
                'in_production': data['in_production'] * product.convert_factor if product.convert_factor != 0 else data['in_production'],
                'in_adjust': data['in_adjust'] * product.convert_factor if product.convert_factor != 0 else data['in_adjust'],
                'in_transfer': data['in_transfer'] * product.convert_factor if product.convert_factor != 0 else data['in_transfer'],
                'in_other': data['in_other'] * product.convert_factor if product.convert_factor != 0 else data['in_other'],
                'in_total': data['in_total'] * product.convert_factor if product.convert_factor != 0 else data['in_total'],
                'out_production': data['out_production'] * product.convert_factor if product.convert_factor != 0 else data['out_production'],
                'out_supplier': data['out_supplier'] * product.convert_factor if product.convert_factor != 0 else data['out_supplier'],
                'out_huy': data['out_huy'] * product.convert_factor if product.convert_factor != 0 else data['out_huy'],
                'out_adjust': data['out_adjust'] * product.convert_factor if product.convert_factor != 0 else data['out_adjust'],
                'out_transfer': data['out_transfer'] * product.convert_factor if product.convert_factor != 0 else data['out_transfer'],
                'out_other': data['out_other'] * product.convert_factor if product.convert_factor != 0 else data['out_other'],
                'out_total': data['out_total'] * product.convert_factor if product.convert_factor != 0 else data['out_total'],
                'end_quantity': data['end_quantity'] * product.convert_factor if product.convert_factor != 0 else data['end_quantity'],

            })
        detailed_data = sorted(detailed_data, key=lambda x: x['warehouse'])
        worksheet['A3'] = "Range Date: " + from_date.strftime('%d/%m/%Y') + " - " + to_date.strftime('%d/%m/%Y')

        row = 6
        for data in detailed_data:
            keys = list(data.keys())
            # Nếu bạn cần thứ tự cụ thể và đang sử dụng phiên bản Python cũ hơn, bạn có thể muốn sắp xếp các khóa:
            # keys = sorted(data.keys())
            for index, key in enumerate(keys, start=1):
                worksheet.cell(row, index, data[key])
            row += 1

        return self.save_excel(workbook, file_name)

        
        
    def action__nxt__ma_hq___sl_qd_hq(self):
        file_name = 'Bao cao NXT theo Ma HQ va SL quy doi HQ.xlsx'
        workbook = self.load_excel(file_name)
        worksheet = workbook.active
        from_date = self.from_date
        to_date = self.to_date + timedelta(days=1)

        transaction_data = self._get_stock_data(from_date, to_date)

        # Biến đổi dữ liệu để dễ dàng hiển thị hoặc trả về
        detailed_data = []
        for (product_id, lot_id, warehouse_id), data in transaction_data.items():
            product = self.env['product.product'].browse(product_id)
            lot = self.env['stock.lot'].browse(lot_id) if lot_id else False
            warehouse = self.env['stock.warehouse'].browse(warehouse_id) if warehouse_id else False
            lot_values = self._get_values_from_split(lot.name if lot else '', ':', 4)
            product_values = self._get_values_from_split(product.name if product else '', ':', 2)
            warehouse_name = warehouse.name if warehouse else ''
            detailed_data.append({
                'warehouse': warehouse_name,
                'he_hang': lot_values[0],
                'order': lot_values[1],
                'mass_hanbai': lot_values[2],
                'bom_add': lot_values[3],
                'product': product_values[0],
                'color': product_values[1],
                'ma_hq': product.customs_code,
                'hsqd': product.convert_factor,
                'begin_quantity': data['begin_quantity'] * product.convert_factor if product.convert_factor != 0 else data['begin_quantity'],
                'in_purchase': data['in_purchase'] * product.convert_factor if product.convert_factor != 0 else data['in_purchase'],
                'in_production': data['in_production'] * product.convert_factor if product.convert_factor != 0 else data['in_production'],
                'in_adjust': data['in_adjust'] * product.convert_factor if product.convert_factor != 0 else data['in_adjust'],
                'in_transfer': data['in_transfer'] * product.convert_factor if product.convert_factor != 0 else data['in_transfer'],
                'in_other': data['in_other'] * product.convert_factor if product.convert_factor != 0 else data['in_other'],
                'in_total': data['in_total'] * product.convert_factor if product.convert_factor != 0 else data['in_total'],
                'out_production': data['out_production'] * product.convert_factor if product.convert_factor != 0 else data['out_production'],
                'out_supplier': data['out_supplier'] * product.convert_factor if product.convert_factor != 0 else data['out_supplier'],
                'out_huy': data['out_huy'] * product.convert_factor if product.convert_factor != 0 else data['out_huy'],
                'out_adjust': data['out_adjust'] * product.convert_factor if product.convert_factor != 0 else data['out_adjust'],
                'out_transfer': data['out_transfer'] * product.convert_factor if product.convert_factor != 0 else data['out_transfer'],
                'out_other': data['out_other'] * product.convert_factor if product.convert_factor != 0 else data['out_other'],
                'out_total': data['out_total'] * product.convert_factor if product.convert_factor != 0 else data['out_total'],
                'end_quantity': data['end_quantity'] * product.convert_factor if product.convert_factor != 0 else data['end_quantity'],

            })
        detailed_data = self._group_data(detailed_data)
        detailed_data = sorted(detailed_data, key=lambda x: x['warehouse'])
        worksheet['A3'] = "Range Date: " + from_date.strftime('%d/%m/%Y') + " - " + to_date.strftime('%d/%m/%Y')

        row = 6
        
        for data in detailed_data:
            keys = list(data.keys())
            # Nếu bạn cần thứ tự cụ thể và đang sử dụng phiên bản Python cũ hơn, bạn có thể muốn sắp xếp các khóa:
            # keys = sorted(data.keys())
            for index, key in enumerate(keys, start=1):
                worksheet.cell(row, index, data[key])
            row += 1

        return self.save_excel(workbook, file_name)

        
        
    def _group_data(self,data):
        grouped_data = {}
        # Xử lý dữ liệu để nhóm và cộng dồn
        for item in data:
            key = (item['warehouse'], item['he_hang'], item['mass_hanbai'], item['bom_add'], item['ma_hq'],item['order'])
            if key not in grouped_data:
                # Nếu nhóm chưa tồn tại, tạo mới với các trường số khởi tạo ban đầu
                grouped_data[key] = {
                    'begin_quantity': item['begin_quantity'],
                    'in_purchase': item['in_purchase'],
                    'in_production': item['in_production'],
                    'in_adjust': item['in_adjust'],
                    'in_transfer': item['in_transfer'],
                    'in_other': item['in_other'],
                    'in_total': item['in_total'],
                    'out_production': item['out_production'],
                    'out_supplier': item['out_supplier'],
                    'out_huy': item['out_huy'],
                    'out_adjust': item['out_adjust'],
                    'out_transfer': item['out_transfer'],
                    'out_other': item['out_other'],
                    'out_total': item['out_total'],
                    'end_quantity': item['end_quantity'],
                }
            else:
                # Nếu nhóm đã tồn tại, cập nhật các trường số
                grouped_data[key]['begin_quantity'] += item['begin_quantity']
                grouped_data[key]['in_purchase'] += item['in_purchase']
                grouped_data[key]['in_production'] += item['in_production']
                grouped_data[key]['in_adjust'] += item['in_adjust']
                grouped_data[key]['in_transfer'] += item['in_transfer']
                grouped_data[key]['in_other'] += item['in_other']
                grouped_data[key]['in_total'] += item['in_total']
                grouped_data[key]['out_production'] += item['out_production']
                grouped_data[key]['out_supplier'] += item['out_supplier']
                grouped_data[key]['out_huy'] += item['out_huy']
                grouped_data[key]['out_adjust'] += item['out_adjust']
                grouped_data[key]['out_transfer'] += item['out_transfer']
                grouped_data[key]['out_other'] += item['out_other']
                grouped_data[key]['out_total'] += item['out_total']
                grouped_data[key]['end_quantity'] += item['end_quantity']
           

        # Để xuất kết quả dễ dàng hơn, chuyển dictionary nhóm thành list của dictionaries
        result = []
        for (warehouse, he_hang, mass_hanbai, bom_add, ma_hq, order), totals in grouped_data.items():
            result.append({
                'warehouse': warehouse,
                'he_hang': he_hang,
                'order': order,
                'mass_hanbai': mass_hanbai,
                'bom_add': bom_add,         
                'ma_hq': ma_hq,
                'begin_quantity': totals['begin_quantity'],
                'in_purchase': totals['in_purchase'],
                'in_production': totals['in_production'],
                'in_adjust': totals['in_adjust'],
                'in_transfer': totals['in_transfer'],
                'in_other': totals['in_other'],
                'in_total': totals['in_total'],
                'out_production': totals['out_production'],
                'out_supplier': totals['out_supplier'],
                'out_huy': totals['out_huy'],
                'out_adjust': totals['out_adjust'],
                'out_transfer': totals['out_transfer'],
                'out_other': totals['out_other'],
                'out_total': totals['out_total'],
                'end_quantity': totals['end_quantity'],
            })
        return result
        
    def _classify_transaction(self, move_line):
        transaction_type = 'internal'
        warehouse_id = False

        if move_line.location_id.usage == 'supplier' and move_line.location_dest_id.usage == 'internal':
            transaction_type = 'in_purchase'
            warehouse_id = move_line.location_dest_id.warehouse_id.id
        elif move_line.location_id.usage == 'production' and move_line.location_dest_id.usage == 'internal':
            transaction_type = 'in_production'
            warehouse_id = move_line.location_dest_id.warehouse_id.id
        elif move_line.location_id.usage == 'inventory' and move_line.location_dest_id.usage == 'internal':
            transaction_type = 'in_adjust'
            warehouse_id = move_line.location_dest_id.warehouse_id.id
        elif move_line.location_id.usage == 'transit' and move_line.location_dest_id.usage == 'internal':
            transaction_type = 'in_transfer'
            warehouse_id = move_line.location_dest_id.warehouse_id.id
        elif move_line.location_id.usage == 'internal' and move_line.location_dest_id.usage == 'production':
            transaction_type = 'out_production'
            warehouse_id = move_line.location_id.warehouse_id.id
        elif move_line.location_id.usage == 'internal' and move_line.location_dest_id.usage == 'supplier':
            transaction_type = 'out_supplier'
            warehouse_id = move_line.location_id.warehouse_id.id
        elif move_line.location_id.usage == 'internal' and move_line.location_dest_id.usage == 'inventory':
            if move_line.location_dest_id.scrap_location:
                transaction_type = 'out_scrap'
            else:
                transaction_type = 'out_adjust'
            warehouse_id = move_line.location_id.warehouse_id.id
        elif move_line.location_id.usage == 'internal' and move_line.location_dest_id.usage == 'transit':
            transaction_type = 'out_transfer'
            warehouse_id = move_line.location_id.warehouse_id.id
        elif move_line.location_id.usage == 'internal' and move_line.location_dest_id.usage != 'internal':
            transaction_type = 'out_other'
            warehouse_id = move_line.location_id.warehouse_id.id
        elif move_line.location_dest_id.usage == 'internal' and move_line.location_id.usage != 'internal':
            transaction_type = 'in_other'
            warehouse_id = move_line.location_dest_id.warehouse_id.id

        return transaction_type, warehouse_id
        
    def _get_stock_data(self,from_date,to_date):
        transaction_data = {}
        stockMoveLines = self.env['stock.move.line'].search([('date', '<', to_date), ('state', '=', 'done')])
        from_date_datetime = datetime(from_date.year, from_date.month, from_date.day)
        for line in stockMoveLines:
            data_key, warehouse_id = self._classify_transaction(line)
            if warehouse_id:
                product_key = (line.product_id.id, line.lot_id.id if line.lot_id else False,warehouse_id)
                
                if product_key not in transaction_data:
                    transaction_data[product_key] = {'begin_quantity': 0, 'in_purchase': 0,'in_production': 0,
                                                 'in_adjust': 0,'in_transfer': 0, 'in_other': 0,
                                                 'in_total': 0, 'out_production': 0,'out_supplier': 0,
                                                 'out_huy': 0, 'out_adjust': 0,'out_transfer': 0,
                                                 'out_other': 0, 'out_total': 0,'end_quantity': 0,
                                                }
                if data_key != 'internal':
                    if line.date > from_date_datetime:
                        transaction_data[product_key][data_key] += line.quantity
                    else:
                        if 'in_' in data_key:
                            transaction_data[product_key]['begin_quantity'] += line.quantity
                        else:
                            transaction_data[product_key]['begin_quantity'] -= line.quantity
        for key, values in transaction_data.items():
            transaction_data[key]['in_total'] = values['in_purchase'] + values['in_production'] + values['in_adjust'] + values['in_transfer'] + values['in_other']
            transaction_data[key]['out_total'] = values['out_production'] + values['out_supplier'] + values['out_huy'] + values['out_adjust'] + values['out_transfer'] + values['out_other']
            transaction_data[key]['end_quantity'] = transaction_data[key]['begin_quantity'] + transaction_data[key]['in_total'] - transaction_data[key]['out_total']
        return transaction_data

    def _get_values_from_split(self,name, delimiter, expected_parts):
        """Trả về một list các giá trị từ chuỗi bị split, bổ sung giá trị rỗng nếu cần."""
        parts = (name or '').split(delimiter)
        return [parts[i] if i < len(parts) else '' for i in range(expected_parts)]

class Product_Template(models.Model):
    _inherit = ['product.template']
    _sql_constraints = [
                ('uniq_name', 'unique(name)', "Name Exiting!"),
    ]
    customs_code = fields.Char(string='Customs Code')
    convert_factor = fields.Float(string='Convert Factor')
    convert_rate = fields.Char(string='Convert Rate')
    name = fields.Char(store='True')
    barcode = fields.Char(store='True', default = 'New')


    @api.model
    def create(self, values):
        if values.get('barcode', 'New') == 'New':
           values['barcode'] = self.env['ir.sequence'].next_by_code('product.template') or 'New'


        res = super().create(values)


        return res

class Stock_Picking(models.Model):
    _inherit = ['stock.picking']
    _sql_constraints = [
                ('uniq_name', 'unique(name)', "Name Exiting!"),
    ]
    name = fields.Char(store='True', readonly=False)


    def _package_move_lines(self):
        quantity_move_line_ids = self.move_line_ids.filtered(
            lambda ml:
                float_compare(ml.quantity, 0.0, precision_rounding=ml.product_uom_id.rounding) > 0 and
                not ml.result_package_id 
        )
        move_line_ids = quantity_move_line_ids.filtered(lambda ml: ml.picked and (ml.lot_name or ml.lot_id))
        return move_line_ids  

class stock_moveline(models.Model):
    _inherit = ['stock.move.line']
    product_id = fields.Many2one('product.product', store='True')


    def print_label(self):
        self = self.sudo()
        printer = self.env.user.printing_printer_id or self.env['printing.printer'].search([('name','like','F110_1')],limit=1)
        for record in self:
            if '/IN/' in record.picking_id.name:
                package = record.result_package_id 
                label = self.env['printing.label.zpl2'].search([('name','=','cuon_nguyen')],limit=1)              
                if label and printer:
                    label.print_label(printer, package)
                    return True
            else:             
                label = self.env['printing.label.zpl2'].search([('name','=','cuon_cat')],limit=1)               
                if label and printer:
                    label.print_label(printer, record)
                    return True
        return False
        
    # def write(self, values):
        # res = super(stock_moveline, self).write(values)
        # if 'quantity' in values or 'qty_done' in values:
            # for record in self:
                # record.move_id.write({'product_uom_qty': record.move_id.quantity})
        # return res
    
    # def create(self, values_list):
        # records = super(stock_moveline, self).create(values_list)
        # for record in records:
            # move = record.move_id
            # move.write({'product_uom_qty': move.product_uom_qty + record.quantity})   
        # return records

class Stock_quantpackage(models.Model):
    _inherit = ['stock.quant.package']
    display_name = fields.Char(string='Display Name', compute='_compute_display_name', store=True)
    side = fields.Selection([('L','Trái'),('R','Phải'),], string='Side')
    dye = fields.Char(string='Dye')


    @api.depends('name', 'side', 'dye')
    def _compute_display_name(self):
        for record in self:
            name = record.name  
            if record.side:
                name = name + '-'+ record.side
            if record.dye:
                name = name + '-'+ record.dye
            record.display_name = name

class stock_move(models.Model):
    _inherit = ['stock.move']
    name = fields.Char(store='True')


    def _action_done(self, cancel_backorder=False):
        moves = super(stock_move,self)._action_done(cancel_backorder=cancel_backorder)
        for m in moves:
            dest_moves = m.move_dest_ids.filtered(lambda x: x.state not in ['done', 'cancel'])
            quantity = m.product_uom_qty
            dest_quantity = sum(l.product_uom_qty for l in dest_moves )
            update_qty = quantity - dest_quantity
            for move in dest_moves:
                if update_qty != 0:
                    qty = move.product_uom_qty + update_qty
                    move.write({'product_uom_qty':qty})
                    update_qty = 0
                    move.picking_id.action_assign()
        return moves
